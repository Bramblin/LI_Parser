import requests
import schedule
from bs4 import BeautifulSoup
import fake_useragent
import openpyxl

SITES_VALUE = 0
STATS_VALUE = 0

def main():

    creat_new_excel_file()
    summers = 0

    schedule.every(5).seconds.do(parsing_LI)
    while True:
        schedule.run_pending()

    parsing_LI()


def parsing_LI(): # функция для парсинга

    url = "https://www.liveinternet.ru/rating/ru/sport/today.tsv?"

    fua = fake_useragent.UserAgent().random
    header = {
        'user-agent': fua
    }

    resource = requests.get(url).text
    global sport_list
    sport_list = resource.split("\t")

    create_list_name_sportsites()
    create_list_values_sportssites()
    edit_values_in_excel()
    creat_set_table()

def create_list_name_sportsites(): # для создания списка сайтов

    STEP_SITES_IN_FILE_FROM_LI = 5
    global sport_sites
    sport_sites = [] # create list sports sites

    for b in sport_list:
        if STEP_SITES_IN_FILE_FROM_LI <= len(sport_list):
            sport_sites.append(sport_list[STEP_SITES_IN_FILE_FROM_LI])
            STEP_SITES_IN_FILE_FROM_LI += 6

def create_list_values_sportssites():
    global stat_list
    stat_list = [] # site's statistics per day
    for word in sport_list:
        if word.isnumeric():
            stat_list.append(int(word))

    STEP_VALUES_SITES_FROM_LI = 2
    global summers
    summers = []
    for i in stat_list:
        if STEP_VALUES_SITES_FROM_LI < len(stat_list):
            summers.append(stat_list[STEP_VALUES_SITES_FROM_LI])
            STEP_VALUES_SITES_FROM_LI += 2

    print()

def creat_new_excel_file():
    book = openpyxl.Workbook()
    book.remove(book.active)
    sheet_1 = book.create_sheet("Сводная таблица")
    sheet_2 = book.create_sheet("Данные")
    sheet_1.column_dimensions['A'].width = 100
    sheet_1.column_dimensions['B'].width = 20
    sheet_2.column_dimensions['A'].width = 100
    sheet_2.column_dimensions['B'].width = 20
    book.save("massive.xlsx")
    book.close()
def edit_values_in_excel():
    '''Функция для загрузки данных после создания файла'''
    global SITES_VALUE
    sites_list_number = SITES_VALUE
    global STATS_VALUE
    stats_list_number = STATS_VALUE

    book = openpyxl.load_workbook("massive.xlsx")
    sheet_1 = book.worksheets[1]


    cells_sites = []  # create list cells in excel file
    for i in sport_sites:  # добавление ячеек
        cells_sites.append(f'A{sites_list_number + 1}')
        sites_list_number +=1
    print(cells_sites)
    SITES_VALUE = sites_list_number

    for cell in cells_sites:  # добавление в ячейки файла названий сайтов
        sheet_1[cell] = sport_sites[cells_sites.index(cell)]

    cells_stats = []  # create list cells in excel file
    for i in summers:  # добавление ячеек
        cells_stats.append(f'B{stats_list_number + 1}')
        stats_list_number += 1
    print(cells_stats)
    STATS_VALUE = stats_list_number

    for cell in cells_stats:  # добавление в ячейки файла статистики сайтов
        sheet_1[cell] = summers[cells_stats.index(cell)]


    book.save("massive.xlsx")
    book.close()

def creat_set_table():
    '''Создаёт таблицу уникальных значений'''
    book = openpyxl.load_workbook("massive.xlsx")
    sheet_1 = book.worksheets[1]

    set_sports_site_table_at_excel = set()
    for i in range (1, sheet_1.max_row + 1):
        set_sports_site_table_at_excel.add(sheet_1[i][0].value)
    print(set_sports_site_table_at_excel)
    sheet_1=book.worksheets[0]

    ROW_TABLE = 1
    for save in set_sports_site_table_at_excel:
        c1 = sheet_1.cell(row=ROW_TABLE, column=1)
        c1.value = save
        ROW_TABLE += 1
    book.save("massive.xlsx")
    book.close()

if __name__ == '__main__':
    main()